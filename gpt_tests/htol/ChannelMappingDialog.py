"""VCT 通道映射配置对话框

提供信号到通道的映射配置功能，支持：
- 表格编辑
- 导入/导出 Excel 和 CSV 格式
- 通道号验证和重复检测
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys
import csv
from typing import List, Dict, Tuple, Callable, Optional

# 添加父目录到路径以导入 Logger
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
import Logger

try:
    from tksheet import Sheet
except ImportError:
    Sheet = None


class ChannelMappingDialog:
    """VCT 通道映射配置对话框 - 使用 tksheet 表格"""
    
    def __init__(self, parent, signals: List[str], vct_converter, log_callback: Callable[[str], None]):
        """初始化对话框
        
        Args:
            parent: 父窗口
            signals: 信号名列表
            vct_converter: VCT 转换器实例
            log_callback: 日志回调函数
        """
        self.signals = signals
        self.vct_converter = vct_converter
        self.log = log_callback
        self.result = False
        
        # 创建顶层窗口
        self.top = tk.Toplevel(parent)
        self.top.title("VCT Channel Mapping")
        self.top.geometry("600x650")
        self.top.transient(parent)
        self.top.grab_set()
        
        # 说明标签
        ttk.Label(self.top, text="Click 'Channel' column to edit (separate multiple channels with comma, e.g. 17,25,33)").pack(pady=10)
        
        # 创建表格区域
        frame_table = ttk.Frame(self.top)
        frame_table.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 准备表格数据
        existing_mapping = vct_converter.get_channel_mapping()
        table_data = []
        for i, signal in enumerate(signals):
            channels_str = ""
            if signal in existing_mapping:
                channels_str = ",".join(str(c) for c in existing_mapping[signal])
            table_data.append([i + 1, signal, channels_str])
        
        # 创建 tksheet 表格
        self.sheet = Sheet(
            frame_table,
            data=table_data,
            headers=["No.", "Signal", "Channel"],
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            show_row_index=False,
            header_font=("Microsoft YaHei", 10, "bold"),
            header_align="w"
        )
        self.sheet.pack(fill="both", expand=True)
        
        # 绑定窗口大小变化事件
        self.top.bind("<Configure>", lambda e: self.sheet.refresh())
        
        # 启用编辑功能
        self.sheet.enable_bindings((
            "single_select",
            "cell_select", 
            "edit_cell",
            "arrowkeys",
            "copy",
            "paste",
            "column_width_resize"
        ))
        
        # 设置列宽
        self.sheet.column_width(column=0, width=60)
        self.sheet.column_width(column=1, width=220)
        self.sheet.column_width(column=2, width=250)
        
        # 设置序号列和信号名列为只读
        self.sheet.readonly_columns(columns=[0, 1], readonly=True)
        
        # 设置交替行颜色
        for i in range(len(signals)):
            if i % 2 == 0:
                self.sheet.highlight_rows(rows=[i], bg="#f5f5f5")
        
        # 按钮区域
        frame_buttons = ttk.Frame(self.top)
        frame_buttons.pack(fill="x", padx=10, pady=10)
        
        # 右侧：确定按钮
        ttk.Button(frame_buttons, text="OK", command=self.on_ok).pack(side="right", padx=5)
        
        # 左侧：导入和导出按钮
        ttk.Button(frame_buttons, text="Import", command=self.on_import).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="Export", command=self.on_export).pack(side="left", padx=5)
    
    def validate_channel_string(self, channel_str: str) -> Tuple[bool, List[int], Optional[str]]:
        """验证通道号字符串
        
        Args:
            channel_str: 通道号字符串
            
        Returns:
            (is_valid, channels_list, error_message)
        """
        if not channel_str or not str(channel_str).strip():
            return (True, [], None)
        
        channel_str = str(channel_str)
        channels = []
        parts = channel_str.split(',')
        
        for part in parts:
            part = part.strip()
            if not part:
                continue
            try:
                channel = int(part)
                if 0 <= channel <= 255:
                    channels.append(channel)
                else:
                    return (False, [], f"Channel {channel} out of range (0-255)")
            except ValueError:
                return (False, [], f"'{part}' is not a valid number")
        
        return (True, channels, None)
    
    def _parse_channel_list(self, channel_str: str) -> List[int]:
        """解析通道号字符串，返回通道号列表
        
        支持格式：
        - 单个数字: "5"
        - 逗号分隔: "1,2,3"
        - 范围: "1-5" (表示 1,2,3,4,5)
        """
        channels = []
        if not channel_str:
            return channels
        
        parts = channel_str.replace(' ', '').split(',')
        for part in parts:
            part = part.strip()
            if not part:
                continue
            
            if '-' in part and not part.startswith('-'):
                # 范围格式 "1-5"
                try:
                    start, end = part.split('-', 1)
                    start_num = int(start)
                    end_num = int(end)
                    channels.extend(range(start_num, end_num + 1))
                except ValueError:
                    pass
            else:
                # 单个数字
                try:
                    channels.append(int(part))
                except ValueError:
                    pass
        
        return channels
    
    def on_ok(self):
        """确定按钮点击"""
        mapping = {}
        errors = []
        empty_signals = []
        
        # 从表格获取数据并验证
        table_data = self.sheet.get_sheet_data()
        
        for row in table_data:
            signal = row[1]
            channel_str = row[2]
            
            is_valid, channels, error_msg = self.validate_channel_string(channel_str)
            
            if not is_valid:
                errors.append(f"Signal '{signal}': {error_msg}")
            elif channels:
                mapping[signal] = channels
            else:
                empty_signals.append(signal)
        
        # 如果有格式错误，弹出提示
        if errors:
            error_text = "Invalid channel format:\n\n" + "\n".join(errors)
            messagebox.showerror("Format Error", error_text, parent=self.top)
            return
        
        # 检测重复通道号
        channel_to_signals: Dict[int, List[str]] = {}
        for signal, channels in mapping.items():
            for ch in channels:
                if ch not in channel_to_signals:
                    channel_to_signals[ch] = []
                channel_to_signals[ch].append(signal)
        
        # 找出重复的通道号
        duplicates = {ch: sigs for ch, sigs in channel_to_signals.items() if len(sigs) > 1}
        if duplicates:
            dup_lines = []
            for ch, sigs in sorted(duplicates.items()):
                dup_lines.append(f"  Channel {ch}: {', '.join(sigs)}")
            error_text = "Duplicate channels found:\n\n" + "\n".join(dup_lines)
            messagebox.showerror("Duplicate Channel", error_text, parent=self.top)
            return
        
        # 如果没有配置任何通道，提示用户
        if not mapping:
            result = messagebox.askyesno(
                "Confirm", 
                "No channel mapping configured.\n\nContinue anyway?",
                parent=self.top
            )
            if not result:
                return
        
        # 如果部分信号没有配置，提示用户
        if mapping and empty_signals:
            empty_count = len(empty_signals)
            configured_count = len(mapping)
            
            result = messagebox.askyesno(
                "Confirm", 
                f"{configured_count} signals mapped.\n"
                f"{empty_count} signals not mapped.\n\n"
                f"Continue anyway?",
                parent=self.top
            )
            if not result:
                return
        
        # 保存映射
        if mapping:
            self.vct_converter.set_channel_mapping(mapping)
            self.log(f"Channel mapping saved:")
            for signal, channels in mapping.items():
                self.log(f"  {signal} -> {channels}")
            self.log(f"Total: {len(mapping)} signals mapped")
        
        self.result = True
        self.top.destroy()
    
    def on_cancel(self):
        """取消按钮点击"""
        self.result = False
        self.top.destroy()
    
    def on_import(self):
        """导入通道映射配置（支持 Excel 和 CSV 格式）"""
        file_path = filedialog.askopenfilename(
            title="Import Channel Mapping",
            filetypes=[
                ("Excel Files", "*.xlsx *.xls"),
                ("CSV Files", "*.csv"),
                ("All Files", "*.*")
            ],
            parent=self.top
        )
        
        if not file_path:
            return
        
        try:
            config: Dict[str, List[int]] = {}
            
            # 根据文件扩展名选择解析方式
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext in ('.xlsx', '.xls'):
                # Excel 格式
                try:
                    import openpyxl
                except ImportError:
                    messagebox.showerror(
                        "Import Failed", 
                        "openpyxl library required.\nRun: pip install openpyxl", 
                        parent=self.top
                    )
                    return
                
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    signal = str(row[0]).strip()
                    channel_value = row[1] if len(row) > 1 and row[1] is not None else ""
                    
                    if signal and channel_value:
                        channel_str = str(channel_value).strip()
                        channels = self._parse_channel_list(channel_str)
                        
                        # 合并同一信号的通道
                        if signal in config:
                            config[signal].extend(channels)
                        else:
                            config[signal] = channels
                
                wb.close()
                
            elif ext == '.csv':
                # CSV 格式
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    next(reader, None)  # 跳过标题行
                    
                    for row in reader:
                        if len(row) < 2:
                            continue
                        signal = row[0].strip()
                        channel_str = row[1].strip()
                        
                        if signal and channel_str:
                            channels = self._parse_channel_list(channel_str)
                            
                            # 合并同一信号的通道
                            if signal in config:
                                config[signal].extend(channels)
                            else:
                                config[signal] = channels
            else:
                messagebox.showerror(
                    "Import Failed", 
                    f"Unsupported format: {ext}\nSupported: .xlsx, .xls, .csv", 
                    parent=self.top
                )
                return
            
            # 去重并排序每个信号的通道列表
            for signal in config:
                config[signal] = sorted(set(config[signal]))
            
            # 更新表格数据
            table_data = self.sheet.get_sheet_data()
            updated_count = 0
            
            for i, row in enumerate(table_data):
                signal = row[1]
                if signal in config:
                    channels = config[signal]
                    channel_str = ",".join(str(c) for c in channels)
                    self.sheet.set_cell_data(i, 2, channel_str)
                    updated_count += 1
            
            self.sheet.refresh()
            self.log(f"Imported from {file_path}")
            self.log(f"Updated {updated_count} signal mappings")
            # messagebox.showinfo(
            #     "导入成功", 
            #     f"已导入 {updated_count} 个信号的通道映射配置", 
            #     parent=self.top
            # )

        except Exception as e:
            Logger.error(f"Import failed: {e}", exc_info=True)
            messagebox.showerror("Import Failed", f"Import failed: {e}", parent=self.top)
    
    def on_export(self):
        """导出通道映射配置（支持 Excel 和 CSV 格式）"""
        file_path = filedialog.asksaveasfilename(
            title="Export Channel Mapping",
            filetypes=[
                ("Excel Files", "*.xlsx"),
                ("CSV Files", "*.csv")
            ],
            defaultextension=".xlsx",
            parent=self.top
        )
        
        if not file_path:
            return
        
        try:
            # 从表格收集数据
            config: List[Tuple[str, str]] = []
            table_data = self.sheet.get_sheet_data()
            
            for row in table_data:
                signal = row[1]
                channel_str = row[2]
                
                if channel_str:
                    is_valid, channels, _ = self.validate_channel_string(channel_str)
                    if is_valid and channels:
                        channel_str_formatted = ",".join(str(c) for c in channels)
                        config.append((signal, channel_str_formatted))
            
            # 根据文件扩展名选择保存方式
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.xlsx':
                # Excel 格式
                try:
                    import openpyxl
                    from openpyxl.styles import Font
                except ImportError:
                    messagebox.showerror(
                        "Export Failed", 
                        "openpyxl library required.\nRun: pip install openpyxl", 
                        parent=self.top
                    )
                    return
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Channel Mapping"
                
                # 写入标题行
                ws['A1'] = "Signal"
                ws['B1'] = "Channel"
                ws['A1'].font = Font(bold=True)
                ws['B1'].font = Font(bold=True)
                
                # 写入数据
                for i, (signal, channel_str) in enumerate(config, start=2):
                    ws.cell(row=i, column=1, value=signal)
                    ws.cell(row=i, column=2, value=channel_str)
                
                # 调整列宽
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
                
                wb.save(file_path)
                
            elif ext == '.csv':
                # CSV 格式
                with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Signal", "Channel"])
                    for signal, channel_str in config:
                        writer.writerow([signal, channel_str])
            else:
                messagebox.showerror(
                    "Export Failed", 
                    f"Unsupported format: {ext}\nSupported: .xlsx, .csv", 
                    parent=self.top
                )
                return
            
            self.log(f"Exported to {file_path}")
            self.log(f"Exported {len(config)} signal mappings")
            messagebox.showinfo(
                "Export Success", 
                f"Exported {len(config)} signal mappings", 
                parent=self.top
            )
            
        except Exception as e:
            Logger.error(f"Export failed: {e}", exc_info=True)
            messagebox.showerror("Export Failed", f"Export failed: {e}", parent=self.top)

